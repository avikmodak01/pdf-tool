"""
All-in-One PDF Management Tool
================================
A Flask web application that provides common PDF operations:
  - Merge multiple PDFs
  - Split a PDF into separate pages
  - Extract text from a PDF
  - Extract tables from a PDF into an Excel file
  - Rotate pages
  - Add a password (encrypt)
  - Remove a password (decrypt)
  - Compress / reduce size (via qpdf)
  - Extract images from a PDF
  - Add a text watermark to every page
  - Convert pages to images (PNG)
  - Show PDF info (page count, metadata)

For a beginner from VBA: Think of this file like a VBA module containing
many subroutines (def functions). Flask turns each function into a
"web endpoint" -- a URL the browser can call when a button is clicked.
"""

import difflib
import io
import os
import re
import shutil
import subprocess
import tempfile
import uuid
import zipfile
from pathlib import Path

# Optional libraries — features degrade gracefully if missing
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    from pdf2docx import Converter as DocxConverter
    HAS_PDF2DOCX = True
except ImportError:
    HAS_PDF2DOCX = False

try:
    import mammoth as _mammoth
    HAS_MAMMOTH = True
except ImportError:
    HAS_MAMMOTH = False

try:
    import weasyprint as _weasyprint
    HAS_WEASYPRINT = True
except Exception:
    HAS_WEASYPRINT = False

try:
    import pytesseract
    from pdf2image import convert_from_path as _pdf2img
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

from flask import (
    Flask,
    render_template,
    request,
    send_file,
    jsonify,
    after_this_request,
)
from werkzeug.utils import secure_filename

from pypdf import PdfReader, PdfWriter
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from PIL import Image
import pypdfium2 as pdfium
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent.resolve()
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB upload cap
ALLOWED_EXTENSIONS = {"pdf"}


def allowed_file(filename: str) -> bool:
    """Check that the uploaded file is a PDF."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_upload(file_storage) -> Path:
    """Save an uploaded file to the uploads folder with a unique name and
    return its path. (Unique name prevents two users overwriting each other.)"""
    safe_name = secure_filename(file_storage.filename)
    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    dest = UPLOAD_DIR / unique_name
    file_storage.save(dest)
    return dest


def cleanup_later(*paths):
    """Register files to be deleted after the response is sent."""
    @after_this_request
    def _cleanup(response):
        for p in paths:
            try:
                if p and Path(p).exists():
                    Path(p).unlink()
            except Exception:
                pass
        return response


def send_and_cleanup(filepath: Path, download_name: str, mimetype: str,
                    extra_to_clean=None):
    """Helper: send a file back to the browser and clean up temp files.

    `extra_to_clean` is a list of extra paths to delete after sending.
    """
    to_clean = [filepath]
    if extra_to_clean:
        to_clean.extend(extra_to_clean)
    cleanup_later(*to_clean)
    return send_file(
        filepath,
        as_attachment=True,
        download_name=download_name,
        mimetype=mimetype,
    )


# ---------------------------------------------------------------------------
# Home page
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# 1. Merge PDFs
# ---------------------------------------------------------------------------
@app.route("/merge", methods=["POST"])
def merge_pdfs():
    files = request.files.getlist("files")
    if not files or len(files) < 2:
        return jsonify({"error": "Please upload at least 2 PDF files."}), 400

    saved_paths = []
    writer = PdfWriter()
    try:
        for f in files:
            if not allowed_file(f.filename):
                return jsonify({"error": f"'{f.filename}' is not a PDF."}), 400
            path = save_upload(f)
            saved_paths.append(path)
            reader = PdfReader(str(path))
            for page in reader.pages:
                writer.add_page(page)

        out_path = OUTPUT_DIR / f"merged_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)

        return send_and_cleanup(
            out_path, "merged.pdf", "application/pdf",
            extra_to_clean=saved_paths,
        )
    except Exception as e:
        for p in saved_paths:
            try: p.unlink()
            except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 2. Split PDF (one PDF per page, returned as a ZIP)
# ---------------------------------------------------------------------------
@app.route("/split", methods=["POST"])
def split_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        tmpdir = Path(tempfile.mkdtemp())
        page_paths = []
        for i, page in enumerate(reader.pages, start=1):
            writer = PdfWriter()
            writer.add_page(page)
            page_path = tmpdir / f"page_{i:03d}.pdf"
            with open(page_path, "wb") as fp:
                writer.write(fp)
            page_paths.append(page_path)

        zip_path = OUTPUT_DIR / f"split_{uuid.uuid4().hex}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in page_paths:
                zf.write(p, arcname=p.name)
        shutil.rmtree(tmpdir, ignore_errors=True)

        return send_and_cleanup(
            zip_path, "split_pages.zip", "application/zip",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 3. Extract a specific page range (like "copy pages 3-7 into a new PDF")
# ---------------------------------------------------------------------------
def parse_page_ranges(spec: str, total_pages: int):
    """Parse a string like '1-3,5,7-9' into a list of 0-indexed page numbers.

    Very similar in spirit to Excel's "Print Pages 1-3, 5, 7-9" dialog.
    """
    pages = []
    spec = spec.replace(" ", "")
    if not spec:
        raise ValueError("Page range cannot be empty.")
    for part in spec.split(","):
        if "-" in part:
            start, end = part.split("-", 1)
            s, e = int(start), int(end)
            if s < 1 or e < 1 or s > total_pages or e > total_pages or s > e:
                raise ValueError(f"Invalid range: {part}")
            pages.extend(range(s - 1, e))
        else:
            n = int(part)
            if n < 1 or n > total_pages:
                raise ValueError(f"Invalid page: {part}")
            pages.append(n - 1)
    return pages


@app.route("/extract-pages", methods=["POST"])
def extract_pages():
    file = request.files.get("file")
    page_spec = request.form.get("pages", "").strip()
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    if not page_spec:
        return jsonify({"error": "Please enter a page range (e.g. 1-3,5)."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        total = len(reader.pages)
        page_indices = parse_page_ranges(page_spec, total)

        writer = PdfWriter()
        for idx in page_indices:
            writer.add_page(reader.pages[idx])

        out_path = OUTPUT_DIR / f"extracted_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)

        return send_and_cleanup(
            out_path, "extracted_pages.pdf", "application/pdf",
            extra_to_clean=[src_path],
        )
    except ValueError as ve:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 4. Extract text from a PDF (as a .txt file)
# ---------------------------------------------------------------------------
@app.route("/extract-text", methods=["POST"])
def extract_text():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    try:
        text_parts = []
        with pdfplumber.open(src_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                page_text = page.extract_text() or ""
                text_parts.append(f"--- Page {i} ---\n{page_text}\n")

        out_path = OUTPUT_DIR / f"text_{uuid.uuid4().hex}.txt"
        out_path.write_text("\n".join(text_parts), encoding="utf-8")
        return send_and_cleanup(
            out_path, "extracted_text.txt", "text/plain",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 5. Extract tables from a PDF into an Excel file
# ---------------------------------------------------------------------------
@app.route("/extract-tables", methods=["POST"])
def extract_tables():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    try:
        wb = Workbook()
        wb.remove(wb.active)  # start empty
        table_count = 0

        with pdfplumber.open(src_path) as pdf:
            for page_idx, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                for t_idx, table in enumerate(tables, start=1):
                    if not table:
                        continue
                    table_count += 1
                    sheet_name = f"P{page_idx}_T{t_idx}"[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    for row in table:
                        ws.append([c if c is not None else "" for c in row])

        if table_count == 0:
            # No tables found -- create a friendly note sheet
            ws = wb.create_sheet(title="No tables")
            ws.append(["No tables were detected in this PDF."])

        out_path = OUTPUT_DIR / f"tables_{uuid.uuid4().hex}.xlsx"
        wb.save(out_path)
        return send_and_cleanup(
            out_path, "extracted_tables.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 6. Rotate all pages by a chosen angle
# ---------------------------------------------------------------------------
@app.route("/rotate", methods=["POST"])
def rotate_pdf():
    file = request.files.get("file")
    angle = request.form.get("angle", "90")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    try:
        angle_int = int(angle)
        if angle_int not in (90, 180, 270, -90):
            return jsonify({"error": "Angle must be 90, 180, or 270."}), 400
    except ValueError:
        return jsonify({"error": "Invalid angle."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        for page in reader.pages:
            page.rotate(angle_int)
            writer.add_page(page)

        out_path = OUTPUT_DIR / f"rotated_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(
            out_path, "rotated.pdf", "application/pdf",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 7. Encrypt (add a password)
# ---------------------------------------------------------------------------
@app.route("/encrypt", methods=["POST"])
def encrypt_pdf():
    file = request.files.get("file")
    password = request.form.get("password", "")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    if not password:
        return jsonify({"error": "Please enter a password."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        writer.encrypt(password)

        out_path = OUTPUT_DIR / f"encrypted_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(
            out_path, "encrypted.pdf", "application/pdf",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 8. Decrypt (remove a password)
# ---------------------------------------------------------------------------
@app.route("/decrypt", methods=["POST"])
def decrypt_pdf():
    file = request.files.get("file")
    password = request.form.get("password", "")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        if reader.is_encrypted:
            result = reader.decrypt(password)
            if not result:
                return jsonify({"error": "Wrong password."}), 400

        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)

        out_path = OUTPUT_DIR / f"decrypted_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(
            out_path, "decrypted.pdf", "application/pdf",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 9. Compress / optimize a PDF (uses the qpdf command-line tool)
# ---------------------------------------------------------------------------
@app.route("/compress", methods=["POST"])
def compress_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    out_path = OUTPUT_DIR / f"compressed_{uuid.uuid4().hex}.pdf"
    try:
        # qpdf with object stream + stream data compression
        result = subprocess.run(
            [
                "qpdf",
                "--linearize",
                "--object-streams=generate",
                "--compress-streams=y",
                "--recompress-flate",
                "--compression-level=9",
                str(src_path),
                str(out_path),
            ],
            capture_output=True, text=True, timeout=180,
        )
        # qpdf may return 3 for warnings but still produce output
        if result.returncode not in (0, 3) or not out_path.exists():
            return jsonify({"error": f"qpdf failed: {result.stderr}"}), 500

        original = src_path.stat().st_size
        compressed = out_path.stat().st_size
        # If result is larger, just return the original
        if compressed >= original:
            shutil.copy(src_path, out_path)

        return send_and_cleanup(
            out_path, "compressed.pdf", "application/pdf",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        try: out_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 10. Add a text watermark on every page
# ---------------------------------------------------------------------------
def make_watermark_pdf(text: str, page_width: float, page_height: float) -> io.BytesIO:
    """Build an in-memory one-page PDF with a diagonal watermark."""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(page_width, page_height))
    c.saveState()
    c.setFillColorRGB(0.7, 0.7, 0.7, alpha=0.35)
    c.setFont("Helvetica-Bold", max(40, int(page_width / 10)))
    c.translate(page_width / 2, page_height / 2)
    c.rotate(45)
    c.drawCentredString(0, 0, text)
    c.restoreState()
    c.save()
    buf.seek(0)
    return buf


@app.route("/watermark", methods=["POST"])
def watermark_pdf():
    file = request.files.get("file")
    text = request.form.get("text", "").strip()
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    if not text:
        return jsonify({"error": "Please enter watermark text."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        for page in reader.pages:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            wm_pdf = make_watermark_pdf(text, w, h)
            wm_page = PdfReader(wm_pdf).pages[0]
            page.merge_page(wm_page)
            writer.add_page(page)

        out_path = OUTPUT_DIR / f"watermarked_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(
            out_path, "watermarked.pdf", "application/pdf",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 11. Convert every page to a PNG image (zipped)
# ---------------------------------------------------------------------------
@app.route("/to-images", methods=["POST"])
def pdf_to_images():
    file = request.files.get("file")
    dpi_str = request.form.get("dpi", "150")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    try:
        dpi = max(72, min(int(dpi_str), 300))
    except ValueError:
        dpi = 150

    src_path = save_upload(file)
    tmpdir = Path(tempfile.mkdtemp())
    try:
        pdf = pdfium.PdfDocument(str(src_path))
        scale = dpi / 72.0
        image_paths = []
        for i in range(len(pdf)):
            page = pdf[i]
            pil_image = page.render(scale=scale).to_pil()
            img_path = tmpdir / f"page_{i+1:03d}.png"
            pil_image.save(img_path, "PNG")
            image_paths.append(img_path)

        zip_path = OUTPUT_DIR / f"images_{uuid.uuid4().hex}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in image_paths:
                zf.write(p, arcname=p.name)
        shutil.rmtree(tmpdir, ignore_errors=True)

        return send_and_cleanup(
            zip_path, "pages_as_images.zip", "application/zip",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        shutil.rmtree(tmpdir, ignore_errors=True)
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 12. Extract embedded images (uses pdfimages from poppler-utils)
# ---------------------------------------------------------------------------
@app.route("/extract-images", methods=["POST"])
def extract_images():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    tmpdir = Path(tempfile.mkdtemp())
    try:
        prefix = tmpdir / "img"
        result = subprocess.run(
            ["pdfimages", "-all", str(src_path), str(prefix)],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            return jsonify({"error": f"pdfimages failed: {result.stderr}"}), 500

        images = sorted(tmpdir.iterdir())
        if not images:
            return jsonify({"error": "No embedded images were found in this PDF."}), 400

        zip_path = OUTPUT_DIR / f"images_{uuid.uuid4().hex}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in images:
                zf.write(p, arcname=p.name)
        shutil.rmtree(tmpdir, ignore_errors=True)
        return send_and_cleanup(
            zip_path, "extracted_images.zip", "application/zip",
            extra_to_clean=[src_path],
        )
    except Exception as e:
        shutil.rmtree(tmpdir, ignore_errors=True)
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 13. Show PDF info (page count, metadata, encryption status)
# ---------------------------------------------------------------------------
@app.route("/info", methods=["POST"])
def pdf_info():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400

    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        meta = reader.metadata or {}
        info = {
            "filename": file.filename,
            "file_size_kb": round(src_path.stat().st_size / 1024, 2),
            "pages": len(reader.pages),
            "encrypted": reader.is_encrypted,
            "title": str(meta.get("/Title", "") or ""),
            "author": str(meta.get("/Author", "") or ""),
            "subject": str(meta.get("/Subject", "") or ""),
            "creator": str(meta.get("/Creator", "") or ""),
            "producer": str(meta.get("/Producer", "") or ""),
        }
        # Page dimensions of page 1
        if len(reader.pages) > 0:
            p = reader.pages[0]
            info["page1_width_pt"] = round(float(p.mediabox.width), 1)
            info["page1_height_pt"] = round(float(p.mediabox.height), 1)

        try: src_path.unlink()
        except: pass
        return jsonify(info)
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 14. Images -> PDF (combine JPG/PNG images into a single PDF)
# ---------------------------------------------------------------------------
@app.route("/images-to-pdf", methods=["POST"])
def images_to_pdf():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Please upload at least one image."}), 400

    saved_paths = []
    try:
        pil_images = []
        for f in files:
            name = f.filename.lower()
            if not (name.endswith(".jpg") or name.endswith(".jpeg") or name.endswith(".png")):
                return jsonify({"error": f"'{f.filename}' is not JPG/PNG."}), 400
            dest = UPLOAD_DIR / f"{uuid.uuid4().hex}_{secure_filename(f.filename)}"
            f.save(dest)
            saved_paths.append(dest)
            img = Image.open(dest).convert("RGB")
            pil_images.append(img)

        if not pil_images:
            return jsonify({"error": "No valid images."}), 400

        out_path = OUTPUT_DIR / f"from_images_{uuid.uuid4().hex}.pdf"
        first, rest = pil_images[0], pil_images[1:]
        first.save(out_path, "PDF", save_all=True, append_images=rest)

        return send_and_cleanup(
            out_path, "from_images.pdf", "application/pdf",
            extra_to_clean=saved_paths,
        )
    except Exception as e:
        for p in saved_paths:
            try: p.unlink()
            except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 15. Reverse page order
# ---------------------------------------------------------------------------
@app.route("/reverse", methods=["POST"])
def reverse_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        for page in reversed(reader.pages):
            writer.add_page(page)
        out_path = OUTPUT_DIR / f"reversed_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "reversed.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 16. Delete specific pages
# ---------------------------------------------------------------------------
@app.route("/delete-pages", methods=["POST"])
def delete_pages():
    file = request.files.get("file")
    page_spec = request.form.get("pages", "").strip()
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    if not page_spec:
        return jsonify({"error": "Please enter page numbers to delete."}), 400
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        total = len(reader.pages)
        to_delete = set(parse_page_ranges(page_spec, total))
        writer = PdfWriter()
        for i, page in enumerate(reader.pages):
            if i not in to_delete:
                writer.add_page(page)
        if len(writer.pages) == 0:
            return jsonify({"error": "Cannot delete all pages from the PDF."}), 400
        out_path = OUTPUT_DIR / f"deleted_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "pages_deleted.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except ValueError as ve:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 17. Reorder pages
# ---------------------------------------------------------------------------
@app.route("/reorder-pages", methods=["POST"])
def reorder_pages():
    file = request.files.get("file")
    order_spec = request.form.get("order", "").strip()
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    if not order_spec:
        return jsonify({"error": "Please enter the new page order."}), 400
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        total = len(reader.pages)
        order = []
        for part in order_spec.replace(" ", "").split(","):
            n = int(part)
            if n < 1 or n > total:
                raise ValueError(f"Page {n} is out of range (1–{total}).")
            order.append(n - 1)
        writer = PdfWriter()
        for idx in order:
            writer.add_page(reader.pages[idx])
        out_path = OUTPUT_DIR / f"reordered_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "reordered.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except ValueError as ve:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 18. Add blank pages
# ---------------------------------------------------------------------------
@app.route("/add-blank-pages", methods=["POST"])
def add_blank_pages():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    try:
        after = int(request.form.get("after_page", "0"))   # 0 = append to end
        count = max(1, min(int(request.form.get("count", "1")), 20))
        reader = PdfReader(str(src_path))
        total = len(reader.pages)
        if after < 0 or after > total:
            raise ValueError(f"'After page' must be 0–{total}.")
        p0 = reader.pages[0]
        pw, ph = float(p0.mediabox.width), float(p0.mediabox.height)
        # Build blank pages in memory
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=(pw, ph))
        for _ in range(count):
            c.showPage()
        c.save()
        buf.seek(0)
        blank_reader = PdfReader(buf)
        writer = PdfWriter()
        for i, page in enumerate(reader.pages):
            writer.add_page(page)
            if after != 0 and i + 1 == after:
                for bp in blank_reader.pages:
                    writer.add_page(bp)
        if after == 0:
            for bp in blank_reader.pages:
                writer.add_page(bp)
        out_path = OUTPUT_DIR / f"blanks_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "with_blank_pages.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except ValueError as ve:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 19. Add page numbers
# ---------------------------------------------------------------------------
def _page_number_overlay(num: int, total: int, w: float, h: float,
                          position: str, fmt: str, font_size: int) -> io.BytesIO:
    labels = {
        "n":              str(num),
        "n_of_total":     f"{num} / {total}",
        "page_n":         f"Page {num}",
        "page_n_of_total": f"Page {num} of {total}",
    }
    text = labels.get(fmt, str(num))
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(w, h))
    c.setFont("Helvetica", font_size)
    c.setFillColorRGB(0.3, 0.3, 0.3)
    mg = 24
    if   position == "bottom-center": c.drawCentredString(w / 2, mg, text)
    elif position == "bottom-right":  c.drawRightString(w - mg, mg, text)
    elif position == "bottom-left":   c.drawString(mg, mg, text)
    elif position == "top-center":    c.drawCentredString(w / 2, h - mg - font_size, text)
    elif position == "top-right":     c.drawRightString(w - mg, h - mg - font_size, text)
    elif position == "top-left":      c.drawString(mg, h - mg - font_size, text)
    c.save()
    buf.seek(0)
    return buf


@app.route("/add-page-numbers", methods=["POST"])
def add_page_numbers():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    position  = request.form.get("position", "bottom-center")
    fmt       = request.form.get("format", "n_of_total")
    start_num = int(request.form.get("start", "1") or "1")
    font_size = int(request.form.get("font_size", "11") or "11")
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        total = len(reader.pages)
        for i, page in enumerate(reader.pages):
            w, h = float(page.mediabox.width), float(page.mediabox.height)
            overlay = _page_number_overlay(
                start_num + i, total + start_num - 1, w, h, position, fmt, font_size)
            page.merge_page(PdfReader(overlay).pages[0])
            writer.add_page(page)
        out_path = OUTPUT_DIR / f"numbered_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "page_numbered.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 20. Add header / footer
# ---------------------------------------------------------------------------
def _header_footer_overlay(header: str, footer: str, w: float, h: float,
                            font_size: int) -> io.BytesIO:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(w, h))
    c.setFont("Helvetica", font_size)
    c.setFillColorRGB(0.3, 0.3, 0.3)
    mg = 20
    if header:
        c.drawCentredString(w / 2, h - mg - font_size, header)
    if footer:
        c.drawCentredString(w / 2, mg, footer)
    c.save()
    buf.seek(0)
    return buf


@app.route("/add-header-footer", methods=["POST"])
def add_header_footer():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    header    = request.form.get("header", "").strip()
    footer    = request.form.get("footer", "").strip()
    font_size = int(request.form.get("font_size", "11") or "11")
    if not header and not footer:
        return jsonify({"error": "Enter at least a header or footer text."}), 400
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        for page in reader.pages:
            w, h = float(page.mediabox.width), float(page.mediabox.height)
            overlay = _header_footer_overlay(header, footer, w, h, font_size)
            page.merge_page(PdfReader(overlay).pages[0])
            writer.add_page(page)
        out_path = OUTPUT_DIR / f"hf_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "header_footer.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 21. Grayscale
# ---------------------------------------------------------------------------
@app.route("/grayscale", methods=["POST"])
def grayscale_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    try:
        pdf_doc = pdfium.PdfDocument(str(src_path))
        images = []
        for i in range(len(pdf_doc)):
            pil_img = pdf_doc[i].render(scale=2.0).to_pil().convert("L").convert("RGB")
            images.append(pil_img)
        out_path = OUTPUT_DIR / f"grayscale_{uuid.uuid4().hex}.pdf"
        images[0].save(str(out_path), "PDF", save_all=True, append_images=images[1:])
        return send_and_cleanup(out_path, "grayscale.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 22. Remove metadata
# ---------------------------------------------------------------------------
@app.route("/remove-metadata", methods=["POST"])
def remove_metadata():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        writer.add_metadata({"/Title": "", "/Author": "", "/Subject": "",
                             "/Creator": "", "/Producer": "", "/Keywords": ""})
        out_path = OUTPUT_DIR / f"clean_{uuid.uuid4().hex}.pdf"
        with open(out_path, "wb") as fp:
            writer.write(fp)
        return send_and_cleanup(out_path, "metadata_removed.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 23. Flatten (remove interactive elements / annotations)
# ---------------------------------------------------------------------------
@app.route("/flatten", methods=["POST"])
def flatten_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    out_path = OUTPUT_DIR / f"flattened_{uuid.uuid4().hex}.pdf"
    try:
        if HAS_FITZ:
            doc = fitz.open(str(src_path))
            for page in doc:
                for annot in list(page.annots()):
                    page.delete_annot(annot)
            doc.save(str(out_path), garbage=4, deflate=True)
            doc.close()
        else:
            # Fallback: render to images (fully flattens, loses text layer)
            pdf_doc = pdfium.PdfDocument(str(src_path))
            imgs = [pdf_doc[i].render(scale=2.0).to_pil().convert("RGB")
                    for i in range(len(pdf_doc))]
            imgs[0].save(str(out_path), "PDF", save_all=True, append_images=imgs[1:])
        return send_and_cleanup(out_path, "flattened.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 24. Redact text
# ---------------------------------------------------------------------------
@app.route("/redact", methods=["POST"])
def redact_pdf():
    file = request.files.get("file")
    search_text = request.form.get("text", "").strip()
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    if not search_text:
        return jsonify({"error": "Please enter the text to redact."}), 400
    if not HAS_FITZ:
        return jsonify({"error": "Redact requires PyMuPDF. Run: pip3 install pymupdf"}), 500
    src_path = save_upload(file)
    try:
        doc = fitz.open(str(src_path))
        count = 0
        for page in doc:
            hits = page.search_for(search_text)
            count += len(hits)
            for rect in hits:
                page.add_redact_annot(rect, fill=(0, 0, 0))
            page.apply_redactions()
        if count == 0:
            doc.close()
            try: src_path.unlink()
            except: pass
            return jsonify({"error": f"'{search_text}' was not found in the PDF."}), 400
        out_path = OUTPUT_DIR / f"redacted_{uuid.uuid4().hex}.pdf"
        doc.save(str(out_path))
        doc.close()
        return send_and_cleanup(out_path, "redacted.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 25. Repair / recover a corrupted PDF
# ---------------------------------------------------------------------------
@app.route("/repair", methods=["POST"])
def repair_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    out_path = OUTPUT_DIR / f"repaired_{uuid.uuid4().hex}.pdf"
    try:
        if HAS_FITZ:
            doc = fitz.open(str(src_path))
            doc.save(str(out_path), garbage=4, deflate=True, clean=True)
            doc.close()
        else:
            reader = PdfReader(str(src_path))
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            with open(out_path, "wb") as fp:
                writer.write(fp)
        return send_and_cleanup(out_path, "repaired.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 26. PDF → Word (.docx)
# ---------------------------------------------------------------------------
@app.route("/pdf-to-word", methods=["POST"])
def pdf_to_word():
    if not HAS_PDF2DOCX:
        return jsonify({"error": "pdf2docx not installed. Run: pip3 install pdf2docx"}), 500
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    out_path = OUTPUT_DIR / f"converted_{uuid.uuid4().hex}.docx"
    try:
        cv = DocxConverter(str(src_path))
        cv.convert(str(out_path))
        cv.close()
        return send_and_cleanup(
            out_path, "converted.docx",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        try: out_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 27. HTML → PDF
# ---------------------------------------------------------------------------
@app.route("/html-to-pdf", methods=["POST"])
def html_to_pdf():
    if not HAS_WEASYPRINT:
        return jsonify({"error": "weasyprint not installed. Run: pip3 install weasyprint"}), 500
    url       = request.form.get("url", "").strip()
    html_file = request.files.get("file")
    out_path  = OUTPUT_DIR / f"from_html_{uuid.uuid4().hex}.pdf"
    src_path  = None
    try:
        if url:
            _weasyprint.HTML(url=url).write_pdf(str(out_path))
        elif html_file:
            src_path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{secure_filename(html_file.filename)}"
            html_file.save(src_path)
            _weasyprint.HTML(filename=str(src_path)).write_pdf(str(out_path))
        else:
            return jsonify({"error": "Provide a URL or upload an HTML file."}), 400
        return send_and_cleanup(out_path, "from_html.pdf", "application/pdf",
                                extra_to_clean=[src_path] if src_path else [])
    except Exception as e:
        if src_path:
            try: src_path.unlink()
            except: pass
        try: out_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 28. Word count
# ---------------------------------------------------------------------------
@app.route("/word-count", methods=["POST"])
def word_count():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    try:
        total_words = total_chars = 0
        pages_data = []
        with pdfplumber.open(src_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text  = page.extract_text() or ""
                words = len(text.split())
                chars = len(text.replace("\n", ""))
                total_words += words
                total_chars += chars
                pages_data.append({"page": i, "words": words, "characters": chars})
        try: src_path.unlink()
        except: pass
        return jsonify({"total_pages": len(pages_data),
                        "total_words": total_words,
                        "total_characters": total_chars,
                        "pages": pages_data})
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 29. Extract hyperlinks
# ---------------------------------------------------------------------------
@app.route("/extract-links", methods=["POST"])
def extract_links():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    src_path = save_upload(file)
    try:
        reader = PdfReader(str(src_path))
        links = []
        for i, page in enumerate(reader.pages, start=1):
            annots = page.get("/Annots")
            if not annots:
                continue
            for annot_ref in annots:
                obj = annot_ref.get_object()
                if obj.get("/Subtype") == "/Link":
                    action = obj.get("/A", {})
                    uri = action.get("/URI")
                    if uri:
                        links.append({"page": i, "url": str(uri)})
        try: src_path.unlink()
        except: pass
        return jsonify({"total_links": len(links), "links": links})
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 30. Compare two PDFs (text diff)
# ---------------------------------------------------------------------------
@app.route("/compare", methods=["POST"])
def compare_pdfs():
    f1 = request.files.get("file1")
    f2 = request.files.get("file2")
    if not f1 or not f2:
        return jsonify({"error": "Please upload two PDF files."}), 400
    if not allowed_file(f1.filename) or not allowed_file(f2.filename):
        return jsonify({"error": "Both files must be PDFs."}), 400
    p1 = save_upload(f1)
    p2 = save_upload(f2)
    try:
        def _lines(path):
            result = []
            with pdfplumber.open(path) as pdf:
                for i, page in enumerate(pdf.pages, start=1):
                    for line in (page.extract_text() or "").splitlines():
                        result.append(f"[Page {i}] {line}")
            return result

        diff = list(difflib.unified_diff(
            _lines(p1), _lines(p2),
            fromfile=f1.filename, tofile=f2.filename, lineterm=""))
        out_path = OUTPUT_DIR / f"diff_{uuid.uuid4().hex}.txt"
        out_path.write_text(
            "\n".join(diff) if diff else "No text differences found between the two PDFs.",
            encoding="utf-8")
        return send_and_cleanup(out_path, "comparison.txt", "text/plain",
                                extra_to_clean=[p1, p2])
    except Exception as e:
        for p in (p1, p2):
            try: p.unlink()
            except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 31. N-up layout (2-up or 4-up pages per sheet)
# ---------------------------------------------------------------------------
@app.route("/nup", methods=["POST"])
def nup_pdf():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    try:
        n = int(request.form.get("layout", "2"))
        if n not in (2, 4):
            raise ValueError()
    except ValueError:
        return jsonify({"error": "Layout must be 2 or 4."}), 400
    src_path = save_upload(file)
    try:
        pdf_doc = pdfium.PdfDocument(str(src_path))
        page_images = [pdf_doc[i].render(scale=1.5).to_pil().convert("RGB")
                       for i in range(len(pdf_doc))]
        cols = 2
        rows = 1 if n == 2 else 2
        sheet_w, sheet_h = 1684, 1190   # landscape ~A3 @ ~144 dpi
        cell_w, cell_h   = sheet_w // cols, sheet_h // rows
        output_images = []
        for start in range(0, len(page_images), n):
            chunk = page_images[start:start + n]
            sheet = Image.new("RGB", (sheet_w, sheet_h), (255, 255, 255))
            for idx, img in enumerate(chunk):
                col, row = idx % cols, idx // cols
                thumb = img.copy()
                thumb.thumbnail((cell_w - 12, cell_h - 12), Image.LANCZOS)
                x = col * cell_w + (cell_w - thumb.width) // 2
                y = row * cell_h + (cell_h - thumb.height) // 2
                sheet.paste(thumb, (x, y))
            output_images.append(sheet)
        out_path = OUTPUT_DIR / f"nup_{uuid.uuid4().hex}.pdf"
        output_images[0].save(str(out_path), "PDF", save_all=True,
                              append_images=output_images[1:])
        return send_and_cleanup(out_path, f"{n}up.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 32. Batch compress
# ---------------------------------------------------------------------------
@app.route("/batch-compress", methods=["POST"])
def batch_compress():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Please upload at least one PDF file."}), 400
    saved, outputs = [], []
    try:
        for f in files:
            if not allowed_file(f.filename):
                return jsonify({"error": f"'{f.filename}' is not a PDF."}), 400
            saved.append(save_upload(f))
        for i, src in enumerate(saved):
            out = OUTPUT_DIR / f"comp_{uuid.uuid4().hex}.pdf"
            res = subprocess.run(
                ["qpdf", "--linearize", "--object-streams=generate",
                 "--compress-streams=y", "--recompress-flate",
                 "--compression-level=9", str(src), str(out)],
                capture_output=True, text=True, timeout=180)
            if res.returncode not in (0, 3) or not out.exists():
                shutil.copy(src, out)
            elif out.stat().st_size >= src.stat().st_size:
                shutil.copy(src, out)
            outputs.append((files[i].filename, out))
        zip_path = OUTPUT_DIR / f"batch_compressed_{uuid.uuid4().hex}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for orig_name, out in outputs:
                stem = Path(orig_name).stem
                zf.write(out, arcname=f"{stem}_compressed.pdf")
        return send_and_cleanup(zip_path, "batch_compressed.zip", "application/zip",
                                extra_to_clean=saved + [o for _, o in outputs])
    except Exception as e:
        for p in saved + [o for _, o in outputs]:
            try: p.unlink()
            except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 33. Batch watermark
# ---------------------------------------------------------------------------
@app.route("/batch-watermark", methods=["POST"])
def batch_watermark():
    files = request.files.getlist("files")
    text  = request.form.get("text", "").strip()
    if not files:
        return jsonify({"error": "Please upload at least one PDF file."}), 400
    if not text:
        return jsonify({"error": "Please enter watermark text."}), 400
    saved, outputs = [], []
    try:
        for f in files:
            if not allowed_file(f.filename):
                return jsonify({"error": f"'{f.filename}' is not a PDF."}), 400
            saved.append(save_upload(f))
        for i, src in enumerate(saved):
            reader = PdfReader(str(src))
            writer = PdfWriter()
            for page in reader.pages:
                w, h = float(page.mediabox.width), float(page.mediabox.height)
                wm_buf  = make_watermark_pdf(text, w, h)
                wm_page = PdfReader(wm_buf).pages[0]
                page.merge_page(wm_page)
                writer.add_page(page)
            out = OUTPUT_DIR / f"wm_{uuid.uuid4().hex}.pdf"
            with open(out, "wb") as fp:
                writer.write(fp)
            outputs.append((files[i].filename, out))
        zip_path = OUTPUT_DIR / f"batch_watermarked_{uuid.uuid4().hex}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for orig_name, out in outputs:
                stem = Path(orig_name).stem
                zf.write(out, arcname=f"{stem}_watermarked.pdf")
        return send_and_cleanup(zip_path, "batch_watermarked.zip", "application/zip",
                                extra_to_clean=saved + [o for _, o in outputs])
    except Exception as e:
        for p in saved + [o for _, o in outputs]:
            try: p.unlink()
            except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 34. Batch convert to images
# ---------------------------------------------------------------------------
@app.route("/batch-to-images", methods=["POST"])
def batch_to_images():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Please upload at least one PDF file."}), 400
    try:
        dpi   = max(72, min(int(request.form.get("dpi", "150") or "150"), 300))
    except ValueError:
        dpi = 150
    saved = []
    try:
        for f in files:
            if not allowed_file(f.filename):
                return jsonify({"error": f"'{f.filename}' is not a PDF."}), 400
            saved.append(save_upload(f))
        scale    = dpi / 72.0
        zip_path = OUTPUT_DIR / f"batch_imgs_{uuid.uuid4().hex}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, src in enumerate(saved):
                stem    = Path(files[i].filename).stem
                pdf_doc = pdfium.PdfDocument(str(src))
                for pg in range(len(pdf_doc)):
                    buf = io.BytesIO()
                    pdf_doc[pg].render(scale=scale).to_pil().save(buf, "PNG")
                    zf.writestr(f"{stem}/page_{pg+1:03d}.png", buf.getvalue())
        return send_and_cleanup(zip_path, "batch_images.zip", "application/zip",
                                extra_to_clean=saved)
    except Exception as e:
        for p in saved:
            try: p.unlink()
            except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 35. OCR (scanned PDF → searchable text or PDF)
# ---------------------------------------------------------------------------
@app.route("/ocr", methods=["POST"])
def ocr_pdf():
    if not HAS_OCR:
        return jsonify({"error": (
            "OCR requires pytesseract + pdf2image. "
            "Run: pip3 install pytesseract pdf2image  "
            "and install Tesseract: brew install tesseract")}), 500
    file = request.files.get("file")
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Please upload a PDF file."}), 400
    lang        = request.form.get("lang", "eng")
    output_type = request.form.get("output", "txt")   # "txt" or "pdf"
    src_path = save_upload(file)
    try:
        images = _pdf2img(str(src_path), dpi=200)
        if output_type == "txt":
            parts = [f"--- Page {i} ---\n{pytesseract.image_to_string(img, lang=lang)}\n"
                     for i, img in enumerate(images, start=1)]
            out_path = OUTPUT_DIR / f"ocr_{uuid.uuid4().hex}.txt"
            out_path.write_text("\n".join(parts), encoding="utf-8")
            return send_and_cleanup(out_path, "ocr_text.txt", "text/plain",
                                    extra_to_clean=[src_path])
        else:
            # Build searchable PDF: one page per image via pytesseract
            writer = PdfWriter()
            for img in images:
                pdf_bytes = pytesseract.image_to_pdf_or_hocr(img, lang=lang, extension="pdf")
                reader = PdfReader(io.BytesIO(pdf_bytes))
                for page in reader.pages:
                    writer.add_page(page)
            out_path = OUTPUT_DIR / f"searchable_{uuid.uuid4().hex}.pdf"
            with open(out_path, "wb") as fp:
                writer.write(fp)
            return send_and_cleanup(out_path, "searchable.pdf", "application/pdf",
                                    extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# 36. Word → PDF
# ---------------------------------------------------------------------------
def _libreoffice_bin():
    """Return the path to the LibreOffice binary, or None if not found."""
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            return path
    mac = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    return mac if Path(mac).exists() else None


def _word_to_pdf_mammoth(src_path: Path, out_path: Path):
    """Fallback: mammoth docx→HTML, then reportlab HTML→simple PDF."""
    with open(src_path, "rb") as fh:
        result = _mammoth.convert_to_html(fh)
    html = result.value

    # Strip tags to extract plain text (reportlab-safe)
    import html as _html_mod
    import re
    text = _html_mod.unescape(re.sub(r"<[^>]+>", "\n", html))
    lines = [l for l in text.splitlines() if l.strip()]

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []
    for line in lines:
        story.append(Paragraph(line, styles["Normal"]))
        story.append(Spacer(1, 4))
    doc.build(story)


@app.route("/word-to-pdf", methods=["POST"])
def word_to_pdf_route():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Please upload a Word file (.docx / .doc / .odt / .rtf)."}), 400

    fname = secure_filename(file.filename)
    allowed_exts = (".docx", ".doc", ".odt", ".rtf")
    if not any(fname.lower().endswith(ext) for ext in allowed_exts):
        return jsonify({"error": "Unsupported file type. Upload .docx, .doc, .odt or .rtf."}), 400

    src_path = UPLOAD_DIR / f"{uuid.uuid4().hex}_{fname}"
    file.save(src_path)

    lo = _libreoffice_bin()

    try:
        if lo:
            # Primary path — LibreOffice (full fidelity)
            res = subprocess.run(
                [lo, "--headless", "--convert-to", "pdf",
                 "--outdir", str(OUTPUT_DIR), str(src_path)],
                capture_output=True, text=True, timeout=180,
            )
            out_path = OUTPUT_DIR / (src_path.stem + ".pdf")
            if res.returncode != 0 or not out_path.exists():
                return jsonify({"error": f"Conversion failed: {res.stderr or 'unknown error'}"}), 500
        elif HAS_MAMMOTH and fname.lower().endswith(".docx"):
            # Fallback — mammoth + reportlab (text only, no rich formatting)
            out_path = OUTPUT_DIR / f"word_{uuid.uuid4().hex}.pdf"
            _word_to_pdf_mammoth(src_path, out_path)
        else:
            return jsonify({"error": (
                "LibreOffice is required for this conversion. "
                "On macOS: brew install libreoffice  "
                "On Linux: apt-get install -y libreoffice-headless libreoffice-writer"
            )}), 500

        return send_and_cleanup(out_path, "converted.pdf", "application/pdf",
                                extra_to_clean=[src_path])
    except Exception as e:
        try: src_path.unlink()
        except: pass
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Run the server
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # debug=True shows errors in the browser -- great while learning.
    # Turn it off (debug=False) when sharing the tool with others.
    app.run(host="127.0.0.1", port=5000, debug=True)
